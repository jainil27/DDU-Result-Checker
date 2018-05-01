import openpyxl
import webbrowser
from time import sleep
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import datetime

def check_res():
    data=openpyxl.load_workbook('./it_data.xlsx','r')
    sheet = data.active
    print("\n\tDDU IT Student Result Checker")
    print("Enter the name of student : ",end=" ")
    name=input()
    name=name.upper()
    name=name.strip()

    flag=0
    for rowNum in range(2, 130):
        DataName = sheet.cell(row=rowNum, column=2).value
        #print(produceName)
        if(DataName.find(name)!=-1):
            flag=1
            Id = sheet.cell(row=rowNum,column=1).value
            Id=Id.strip()
            Dob = sheet.cell(row=rowNum,column=3).value
            if(isinstance(Dob,datetime.datetime)):
                Dob=Dob.strftime('%m/%d/%Y')
            Dob=Dob.strip()
            d = Dob.split('/')
            if(len(d[0])==1):
                d[0]='0'+d[0]
            if (len(d[1]) == 1):
                d[1] = '0' + d[1]
            if (len(d[2]) == 2):
                d[2] = '19' + d[2]
            Dob='/'.join(d)
            print("Searching Result...")
            print()
            print("Name : "+DataName)
            print("Id   : "+Id)
            print("Dob  : "+Dob)
            break
    if(flag==0):
        print("No such student exists...")
    if(flag):
        sleep(2)
        chrome_options = Options()
        chrome_options.add_argument("disable-infobars")
        driver= webdriver.Chrome(executable_path="C:/Users/Admin/PycharmProjects/Results/chromedriver.exe",chrome_options=chrome_options)
        driver.get("https://egov.ddit.ac.in/index.php?r=site/login")
        uid = driver.find_element_by_id("LoginForm_username")
        pswd = driver.find_element_by_id("LoginForm_password")
        uid.send_keys(Id)
        pswd.send_keys(Dob)
        #print(driver.current_url)
        while(driver.current_url!="https://egov.ddit.ac.in/index.php?r=studentInformation/studentInfo"):
            sleep(2)
        driver.get("https://egov.ddit.ac.in/index.php?r=tblstudentmst/academicHistory")
        sleep(1)
        driver.find_element_by_id('yt10').click()
        print()
while(1):
    check_res()